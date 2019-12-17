import openpyxl
import pickle

wb = openpyxl.load_workbook('AOA.xlsx')
word_aoa = {}

sheet = wb.get_sheet_by_name('Sheet1')

for i in range(1, sheet.max_row):
	word = sheet.cell(row=i, column=1).value
	aoa = sheet.cell(row=i, column=11).value
	if isinstance(aoa, float):
		word_aoa[word] = aoa

with open('aoa.pickle', 'wb') as handle:
    pickle.dump(word_aoa, handle)

